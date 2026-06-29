from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import os
from flask import send_file
import io
from flask import Flask, render_template, request, redirect, url_for, session
import pymysql

from config import *

app = Flask(__name__)
app.secret_key = "chrono_2026"

# =====================================
# REGISTRAR HISTORIAL
# =====================================

def registrar_historial(conexion, id_usuario, modulo, accion, descripcion):

    cursor = conexion.cursor()

    cursor.execute("""

        INSERT INTO historial_cambios
        (
            id_usuario,
            modulo,
            accion,
            descripcion
        )
        VALUES
        (%s,%s,%s,%s)

    """,
    (
        id_usuario,
        modulo,
        accion,
        descripcion
    ))

# =====================================
# MOTOR DE VALIDACIÓN DE HORARIOS
# =====================================

def validar_horario(

    conexion,

    id_docente,

    id_aula,

    id_grupo,

    dia,

    hora_inicio,

    hora_fin,

    cupo,

    id_horario=None

):

    cursor = conexion.cursor()

    # =====================================
    # VALIDAR DOCENTE
    # =====================================

    if id_horario:

        cursor.execute("""

            SELECT id_horario

            FROM horarios

            WHERE id_docente=%s
            AND dia=%s
            AND id_horario<>%s
            AND (
                hora_inicio < %s
                AND hora_fin > %s
            )

        """,
        (
            id_docente,
            dia,
            id_horario,
            hora_fin,
            hora_inicio
        ))

    else:

        cursor.execute("""

            SELECT id_horario

            FROM horarios

            WHERE id_docente=%s
            AND dia=%s
            AND (
                hora_inicio < %s
                AND hora_fin > %s
            )

        """,
        (
            id_docente,
            dia,
            hora_fin,
            hora_inicio
        ))

    if cursor.fetchone():

        return False, "El docente ya tiene asignada otra clase en ese horario."

    # =====================================
    # VALIDAR AULA
    # =====================================

    if id_horario:

        cursor.execute("""

            SELECT id_horario

            FROM horarios

            WHERE id_aula=%s
            AND dia=%s
            AND id_horario<>%s
            AND (
                hora_inicio < %s
                AND hora_fin > %s
            )

        """,
        (
            id_aula,
            dia,
            id_horario,
            hora_fin,
            hora_inicio
        ))

    else:

        cursor.execute("""

            SELECT id_horario

            FROM horarios

            WHERE id_aula=%s
            AND dia=%s
            AND (
                hora_inicio < %s
                AND hora_fin > %s
            )

        """,
        (
            id_aula,
            dia,
            hora_fin,
            hora_inicio
        ))

    if cursor.fetchone():

        return False, "El aula ya se encuentra ocupada."
    
    # =====================================
    # VALIDAR GRUPO
    # =====================================

    if id_horario:

        cursor.execute("""

            SELECT id_horario

            FROM horarios

            WHERE id_grupo=%s
            AND dia=%s
            AND id_horario<>%s
            AND (
                hora_inicio < %s
                AND hora_fin > %s
            )

        """,
        (
            id_grupo,
            dia,
            id_horario,
            hora_fin,
            hora_inicio
        ))

    else:

        cursor.execute("""

            SELECT id_horario

            FROM horarios

            WHERE id_grupo=%s
            AND dia=%s
            AND (
                hora_inicio < %s
                AND hora_fin > %s
            )

        """,
        (
            id_grupo,
            dia,
            hora_fin,
            hora_inicio
        ))

    if cursor.fetchone():

        return False, "El grupo ya tiene una asignación en ese horario."

    # =====================================
    # VALIDAR CUPO DEL AULA
    # =====================================

    cursor.execute("""

    SELECT capacidad

    FROM aulas

    WHERE id_aula=%s

    """,
    (id_aula,))

    aula = cursor.fetchone()

    if aula:

        capacidad = int(aula["capacidad"])

        if int(cupo) > capacidad:

            return False, (
                f"El cupo del grupo ({cupo}) supera la capacidad "
                f"del aula ({capacidad})."
            )

    # =====================================
    # VALIDAR TURNO DEL GRUPO
    # =====================================

    cursor.execute("""

    SELECT turno

    FROM grupos

    WHERE id_grupo=%s

    """,
    (id_grupo,))

    grupo = cursor.fetchone()

    if grupo:

        turno = grupo["turno"]

        hora = int(hora_inicio.split(":")[0])

        if turno == "MATUTINO" and hora >= 14:

            return False, "El grupo pertenece al turno matutino."

        if turno == "VESPERTINO" and hora < 14:

            return False, "El grupo pertenece al turno vespertino."
    
        # =====================================
    # VALIDAR DURACIÓN DE LA CLASE
    # =====================================

    from datetime import datetime

    formato = "%H:%M"

    inicio = datetime.strptime(str(hora_inicio)[:5], formato)

    fin = datetime.strptime(str(hora_fin)[:5], formato)

    if fin <= inicio:

        return False, "La hora de fin debe ser mayor que la hora de inicio."

    duracion = (fin - inicio).seconds / 3600

    if duracion < 1:

        return False, "La duración mínima de una clase es de 1 hora."

    if duracion > 4:

        return False, "La duración máxima permitida es de 4 horas."

    # =====================================
    # TODAS LAS VALIDACIONES FUERON EXITOSAS
    # =====================================

    return True, "Horario válido."


#=====================================
# LOGIN
# =====================================

@app.route("/", methods=["GET", "POST"])

def login():

    if request.method == "POST":

        correo = request.form["correo"]

        password = request.form["password"]

        try:

            conexion = pymysql.connect(

                host=DB_HOST,

                user=DB_USER,

                password=DB_PASSWORD,

                database=DB_NAME

            )

            cursor = conexion.cursor()

            sql = """

            SELECT

                u.id_usuario,

                u.nombre,

                r.nombre

            FROM usuarios u

            INNER JOIN roles r

                ON u.id_rol = r.id_rol

            WHERE u.correo = %s

            AND u.password = %s

            """

            cursor.execute(sql, (correo, password))

            usuario = cursor.fetchone()

            conexion.close()

            if usuario:

                session["usuario_id"] = usuario[0]

                session["nombre"] = usuario[1]

                session["rol"] = usuario[2]

                rol = usuario[2]

                if rol == "Administrador":

                    return redirect(url_for("admin"))

                elif rol == "Planeador":

                    return redirect(url_for("planeador"))

                elif rol == "Coordinador":

                    return redirect(url_for("coordinador"))

                elif rol == "Docente":

                    return redirect(url_for("docente"))

            return "Usuario o contraseña incorrectos"

        except Exception as e:

            return f"Error: {str(e)}"

    return render_template("login.html") 

# =====================================
# DASHBOARD ADMINISTRADOR
# =====================================

@app.route("/admin")
def admin():
    return render_template("admin/dashboard.html")


# =====================================
# PLANEADOR
# =====================================

@app.route("/planeador")
def planeador():

    return render_template(

        "planeador/inicio_plan.html"

    )


# =====================================
# GESTIÓN DE HORARIOS
# =====================================

@app.route("/gestion_horarios", methods=["GET", "POST"])
def gestion_horarios():

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    mensaje = ""
    error = False

    # =====================================
    # REGISTRAR HORARIO
    # =====================================

    if request.method == "POST":

        id_periodo = request.form["periodo"]
        id_grupo = request.form["grupo"]
        id_materia = request.form["materia"]
        id_docente = request.form["docente"]
        id_aula = request.form["aula"]

        dia = request.form["dia"]

        hora_inicio = request.form["hora_inicio"]
        hora_fin = request.form["hora_fin"]

        cupo = request.form["cupo"]

        # =====================================
        # MOTOR DE VALIDACIÓN
        # =====================================

        valido, mensaje = validar_horario(

            conexion,

            id_docente,

            id_aula,

            id_grupo,

            dia,

            hora_inicio,

            hora_fin,

            cupo

        )

        if valido:

                        # =====================================
            # GUARDAR HORARIO
            # =====================================

            cursor.execute("""

                INSERT INTO horarios
                (

                    id_periodo,
                    id_grupo,
                    id_docente,
                    id_materia,
                    id_aula,
                    dia,
                    hora_inicio,
                    hora_fin,
                    cupo,
                    estado

                )

                VALUES
                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)

            """,
            (

                id_periodo,
                id_grupo,
                id_docente,
                id_materia,
                id_aula,
                dia,
                hora_inicio,
                hora_fin,
                cupo,
                "Pendiente"

            ))

            registrar_historial(

                conexion,

                session["usuario_id"],

                "Horarios",

                "Registrar Horario",

                "Generó un horario preliminar."

            )

            conexion.commit()

            mensaje = "Horario registrado correctamente."

            error = False

        else:

            error = True
                # =====================================
    # PERIODOS ACADÉMICOS
    # =====================================

    cursor.execute("""

        SELECT

            id_periodo,
            clave,
            nombre

        FROM periodos_academicos

        WHERE estado='Activo'

        ORDER BY fecha_inicio

    """)

    periodos = cursor.fetchall()

    # =====================================
    # GRUPOS
    # =====================================

    cursor.execute("""

        SELECT

            id_grupo,
            clave_grupo

        FROM grupos

        WHERE activo=1

        ORDER BY semestre,
                 clave_grupo

    """)

    grupos = cursor.fetchall()

    # =====================================
    # MATERIAS
    # =====================================

    cursor.execute("""

        SELECT

            id_materia,
            nombre

        FROM materias

        ORDER BY nombre

    """)

    materias = cursor.fetchall()

    # =====================================
    # DOCENTES
    # =====================================

    cursor.execute("""

        SELECT

            d.id_docente,

            u.nombre

        FROM docentes d

        INNER JOIN usuarios u

            ON d.id_usuario=u.id_usuario

        WHERE d.activo=1

        ORDER BY u.nombre

    """)

    docentes = cursor.fetchall()

    # =====================================
    # AULAS
    # =====================================

    cursor.execute("""

        SELECT

            id_aula,
            aula

        FROM aulas

        ORDER BY aula

    """)

    aulas = cursor.fetchall()

    # =====================================
    # HORARIOS PRELIMINARES
    # =====================================

    cursor.execute("""

        SELECT

            h.id_horario,

            CONCAT(p.clave,' | ',p.nombre) AS periodo,

            g.clave_grupo AS grupo,

            m.nombre AS materia,

            u.nombre AS docente,

            CONCAT(a.aula,' - ',a.edificio) AS aula,

            h.dia,

            TIME_FORMAT(h.hora_inicio,'%H:%i') AS hora_inicio,

            TIME_FORMAT(h.hora_fin,'%H:%i') AS hora_fin,

            h.estado

        FROM horarios h

        INNER JOIN periodos_academicos p

            ON h.id_periodo=p.id_periodo

        INNER JOIN grupos g

            ON h.id_grupo=g.id_grupo

        INNER JOIN materias m

            ON h.id_materia=m.id_materia

        INNER JOIN docentes d

            ON h.id_docente=d.id_docente

        INNER JOIN usuarios u

            ON d.id_usuario=u.id_usuario

        INNER JOIN aulas a

            ON h.id_aula=a.id_aula

        ORDER BY

            p.id_periodo,

            g.clave_grupo,

            h.dia,

            h.hora_inicio

    """)

    horarios = cursor.fetchall()
    
    
    conexion.close()

    return render_template(

        "planeador/gestion_horarios.html",

        periodos=periodos,

        grupos=grupos,

        materias=materias,

        docentes=docentes,

        aulas=aulas,

        horarios=horarios,

        mensaje=mensaje,

        error=error

    )

# =====================================
# VISTA PRELIMINAR
# =====================================

@app.route("/vista_preliminar", methods=["GET", "POST"])
def vista_preliminar():

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    # =====================================
    # PERIODOS
    # =====================================

    cursor.execute("""

        SELECT

            id_periodo,
            clave,
            nombre

        FROM periodos_academicos

        WHERE estado='Activo'

        ORDER BY fecha_inicio

    """)

    periodos = cursor.fetchall()

    # =====================================
    # GRUPOS
    # =====================================

    cursor.execute("""

        SELECT

            id_grupo,
            clave_grupo

        FROM grupos

        WHERE activo=1

        ORDER BY semestre,
                 clave_grupo

    """)

    grupos = cursor.fetchall()

    matriz = []

    if request.method == "POST":

        id_periodo = request.form["periodo"]

        id_grupo = request.form["grupo"]

        cursor.execute("""

            SELECT

                m.nombre AS materia,

                CONCAT(a.aula,' - ',a.edificio) AS aula,

                h.dia,

                TIME_FORMAT(h.hora_inicio,'%%H:%%i') AS hora_inicio,

                TIME_FORMAT(h.hora_fin,'%%H:%%i') AS hora_fin

            FROM horarios h

            INNER JOIN materias m

                ON h.id_materia=m.id_materia

            INNER JOIN aulas a

                ON h.id_aula=a.id_aula

            WHERE

                h.id_periodo=%s

            AND

                h.id_grupo=%s

        """,
        (

            id_periodo,

            id_grupo

        ))

        horarios = cursor.fetchall()

                # =====================================
        # HORAS A MOSTRAR
        # =====================================

        horas = [

            ("07:00", "07:00 - 08:00"),
            ("08:00", "08:00 - 09:00"),
            ("09:00", "09:00 - 10:00"),
            ("10:00", "10:00 - 11:00"),
            ("11:00", "11:00 - 12:00"),
            ("12:00", "12:00 - 13:00"),
            ("13:00", "13:00 - 14:00"),
            ("14:00", "14:00 - 15:00"),
            ("15:00", "15:00 - 16:00"),
            ("16:00", "16:00 - 17:00"),
            ("17:00", "17:00 - 18:00"),
            ("18:00", "18:00 - 19:00"),
            ("19:00", "19:00 - 20:00"),
            ("20:00", "20:00 - 21:00"),
            ("21:00", "21:00 - 22:00")

        ]

        for hora_bd, hora_mostrar in horas:

            fila = {

                "hora": hora_mostrar,

                "lunes": "",

                "martes": "",

                "miercoles": "",

                "jueves": "",

                "viernes": ""

            }

            from datetime import datetime

            hora_actual = datetime.strptime(hora_bd, "%H:%M")

            for h in horarios:

                inicio = datetime.strptime(str(h["hora_inicio"])[:5], "%H:%M")

                fin = datetime.strptime(str(h["hora_fin"])[:5], "%H:%M")

                if inicio <= hora_actual < fin:

                    informacion = f"""

                    <div class="p-2 rounded bg-info text-white">

                        <strong>{h['materia']}</strong><br>

                        Aula {h['aula']}<br>

                        {str(h['hora_inicio'])[:5]} - {str(h['hora_fin'])[:5]}

                    </div>

                    """

                    dia = h["dia"].lower()

                    if dia == "lunes":

                        fila["lunes"] = informacion

                    elif dia == "martes":

                        fila["martes"] = informacion

                    elif dia in ("miércoles", "miercoles"):

                        fila["miercoles"] = informacion

                    elif dia == "jueves":

                        fila["jueves"] = informacion

                    elif dia == "viernes":

                        fila["viernes"] = informacion

            matriz.append(fila)

    conexion.close()

    return render_template(

        "planeador/vista_preliminar.html",

        periodos=periodos,

        grupos=grupos,

        matriz=matriz

    )

# =====================================
# COORDINADOR
# =====================================

@app.route("/coordinador")
def coordinador():

    return render_template(

        "coordinador/inicio_cor.html"

    )


# =====================================
# MODIFICAR HORARIOS
# =====================================

@app.route("/modificar_horarios")
def modificar_horarios():

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    cursor.execute("""

        SELECT

            h.id_horario,

            CONCAT(p.clave,' | ',p.nombre) AS periodo,

            g.clave_grupo AS grupo,

            m.nombre AS materia,

            u.nombre AS docente,

            CONCAT(a.aula, ' - ', a.edificio) AS aula,

            h.dia,

            h.hora_inicio,

            h.hora_fin,

            h.estado

        FROM horarios h

        INNER JOIN periodos_academicos p

            ON h.id_periodo=p.id_periodo

        INNER JOIN grupos g

            ON h.id_grupo=g.id_grupo

        INNER JOIN materias m

            ON h.id_materia=m.id_materia

        INNER JOIN docentes d

            ON h.id_docente=d.id_docente

        INNER JOIN usuarios u

            ON d.id_usuario=u.id_usuario

        INNER JOIN aulas a

            ON h.id_aula=a.id_aula

        ORDER BY

            g.clave_grupo,

            h.dia,

            h.hora_inicio

    """)

    horarios = cursor.fetchall()

    conexion.close()

    return render_template(

        "coordinador/modificar_horarios.html",

        horarios=horarios

    )

# =====================================
# EDITAR HORARIO
# =====================================

@app.route("/editar_horario/<int:id>", methods=["GET", "POST"])
def editar_horario(id):

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    if request.method == "POST":

        id_docente = request.form["docente"]
        id_aula = request.form["aula"]

        dia = request.form["dia"]

        hora_inicio = request.form["hora_inicio"]
        hora_fin = request.form["hora_fin"]

        cupo = request.form["cupo"]

        # Obtener grupo para validación

        cursor.execute("""

            SELECT

                id_grupo

            FROM horarios

            WHERE id_horario=%s

        """,(id,))

        horario_actual = cursor.fetchone()

        valido, mensaje = validar_horario(

            conexion,

            id_docente,

            id_aula,

            horario_actual["id_grupo"],

            dia,

            hora_inicio,

            hora_fin,

            cupo,

            id

        )

        if valido:

            # ===============================
            # RESPALDO EN HISTORIAL
            # ===============================

            registrar_historial(

                conexion,

                session["usuario_id"],

                "Horarios",

                "Modificar Horario",

                f"Modificó el horario {id}."

            )

            cursor.execute("""

                UPDATE horarios

                SET

                    id_docente=%s,

                    id_aula=%s,

                    dia=%s,

                    hora_inicio=%s,

                    hora_fin=%s,

                    estado='Modificado'

                WHERE id_horario=%s

            """,
            (

                id_docente,

                id_aula,

                dia,

                hora_inicio,

                hora_fin,

                id

            ))

            conexion.commit()

            conexion.close()

            return redirect("/modificar_horarios")

        conexion.close()

        return f"""

        <script>

        alert("{mensaje}");

        window.history.back();

        </script>

        """
    
    cursor.execute("""
        SELECT *
        FROM horarios
        WHERE id_horario=%s
    """,(id,))

    horario = cursor.fetchone()

    cursor.execute("""

        SELECT

            d.id_docente,

            u.nombre

        FROM docentes d

        INNER JOIN usuarios u

            ON d.id_usuario = u.id_usuario

        WHERE d.activo = 1

        ORDER BY u.nombre

    """)

    docentes = cursor.fetchall()

    cursor.execute("""

        SELECT

            id_aula,

            aula,

            capacidad

        FROM aulas

        ORDER BY aula

    """)

    aulas = cursor.fetchall()

    conexion.close()

    return render_template(
        "coordinador/editar_horario.html",
        horario=horario,
        docentes=docentes,
        aulas=aulas
    )


@app.route("/excepciones")
def excepciones():

    return render_template(

        "coordinador/excepciones.html"

    )


@app.route("/notificaciones")
def notificaciones():

    return render_template(

        "coordinador/notificaciones.html"

    )


@app.route("/reportes_coordinador")
def reportes_coordinador():

    return render_template(

        "coordinador/reportes.html"

    )


# =====================================
# DOCENTE
# =====================================

@app.route("/docente")
def docente():

    return render_template(

        "docente/inicio_doc.html"

    )


# =====================================
# MI HORARIO DOCENTE
# =====================================

@app.route("/mi_horario")
def mi_horario():

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    # ============================
    # DOCENTE LOGUEADO
    # ============================

    cursor.execute("""

        SELECT id_docente

        FROM docentes

        WHERE id_usuario=%s

    """, (session["usuario_id"],))

    docente = cursor.fetchone()

    if not docente:

        conexion.close()

        return "No existe un docente asociado."

    id_docente = docente["id_docente"]

    # ============================
    # HORARIOS DEL DOCENTE
    # ============================

    cursor.execute("""

        SELECT

            m.nombre AS materia,

            CONCAT(a.aula,' - ',a.edificio) AS aula,

            h.dia,

            h.hora_inicio,

            h.hora_fin

        FROM horarios h

        INNER JOIN materias m
            ON h.id_materia=m.id_materia

        INNER JOIN aulas a
            ON h.id_aula=a.id_aula

        WHERE

            h.id_docente=%s

        AND

            h.estado IN ('Aprobado','Modificado')

        ORDER BY

            h.dia,
            h.hora_inicio

    """,(id_docente,))

    horarios = cursor.fetchall()

    horas = [

        ("07","07 - 08"),
        ("08","08 - 09"),
        ("09","09 - 10"),
        ("10","10 - 11"),
        ("11","11 - 12"),
        ("12","12 - 13"),
        ("13","13 - 14"),
        ("14","14 - 15"),
        ("15","15 - 16"),
        ("16","16 - 17"),
        ("17","17 - 18"),
        ("18","18 - 19"),
        ("19","19 - 20"),
        ("20","20 - 21"),
        ("21","21 - 22")

    ]

    matriz=[]

    for hora_bd,hora_mostrar in horas:

        fila={

            "hora":hora_mostrar,

            "lunes":"",

            "martes":"",

            "miercoles":"",

            "jueves":"",

            "viernes":""

        }

        for h in horarios:

            hora_inicio=str(h["hora_inicio"])
            hora_fin=str(h["hora_fin"])

            hora_inicio_num=hora_inicio.split(":")[0]
            hora_fin_num=hora_fin.split(":")[0]

            if int(hora_inicio_num) <= int(hora_bd) < int(hora_fin_num):

                informacion=f"""

                <div class="bg-info p-2 rounded text-white">

                    <strong>{h['materia']}</strong><br>

                    Aula {h['aula']}<br>

                    {hora_inicio[:5]} - {hora_fin[:5]}

                </div>

                """

                dia=h["dia"].lower()

                if dia=="lunes":

                    fila["lunes"]=informacion

                elif dia=="martes":

                    fila["martes"]=informacion

                elif dia=="miércoles" or dia=="miercoles":

                    fila["miercoles"]=informacion

                elif dia=="jueves":

                    fila["jueves"]=informacion

                elif dia=="viernes":

                    fila["viernes"]=informacion

        matriz.append(fila)

    conexion.close()

    return render_template(

        "docente/mi_horario.html",

        matriz=matriz

    )

# =====================================
# GESTIÓN DE DOCENTES
# =====================================

@app.route("/docentes")
def docentes():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

    cursor = conexion.cursor()

    cursor.execute("""

        SELECT

            d.id_docente,

            d.numero_empleado,

            d.turno,

            d.activo,

            u.nombre,

            u.correo

        FROM docentes d

        INNER JOIN usuarios u

            ON d.id_usuario = u.id_usuario

        ORDER BY u.nombre

    """)

    docentes = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/docentes.html",
        docentes=docentes
    )


# =====================================
# EDITAR DOCENTE
# =====================================

@app.route("/editar_docente/<int:id>", methods=["GET", "POST"])
def editar_docente(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        numero_empleado = request.form["numero_empleado"]
        turno = request.form["turno"]
        activo = request.form["activo"]

        # =============================
        # DISPONIBILIDAD
        # =============================

        lunes = 1 if "lunes" in request.form else 0
        martes = 1 if "martes" in request.form else 0
        miercoles = 1 if "miercoles" in request.form else 0
        jueves = 1 if "jueves" in request.form else 0
        viernes = 1 if "viernes" in request.form else 0

        # =============================
        # ACTUALIZAR DOCENTE
        # =============================

        cursor.execute("""

            UPDATE docentes

            SET

                numero_empleado=%s,

                turno=%s,

                activo=%s

            WHERE id_docente=%s

        """,
        (

            numero_empleado,

            turno,

            activo,

            id

        ))

        # =============================
        # VERIFICAR DISPONIBILIDAD
        # =============================

        cursor.execute("""

            SELECT id_disponibilidad

            FROM disponibilidad_docente

            WHERE id_docente=%s

        """, (id,))

        existe = cursor.fetchone()

        if existe:

            cursor.execute("""

                UPDATE disponibilidad_docente

                SET

                    lunes=%s,

                    martes=%s,

                    miercoles=%s,

                    jueves=%s,

                    viernes=%s

                WHERE id_docente=%s

            """,
            (

                lunes,

                martes,

                miercoles,

                jueves,

                viernes,

                id

            ))

        else:

            cursor.execute("""

                INSERT INTO disponibilidad_docente

                (

                    id_docente,

                    lunes,

                    martes,

                    miercoles,

                    jueves,

                    viernes

                )

                VALUES

                (%s,%s,%s,%s,%s,%s)

            """,
            (

                id,

                lunes,

                martes,

                miercoles,

                jueves,

                viernes

            ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        cursor.execute("""

            SELECT

                u.nombre

            FROM docentes d

            INNER JOIN usuarios u

                ON d.id_usuario = u.id_usuario

            WHERE d.id_docente=%s

        """,(id,))

        docente = cursor.fetchone()

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Docentes",

            "Editar Docente",

            f"Actualizó la información académica del docente '{docente['nombre']}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/docentes")

    # =============================
    # CONSULTA DEL DOCENTE
    # =============================

    cursor.execute("""

        SELECT

            d.id_docente,

            d.numero_empleado,

            d.turno,

            d.activo,

            u.nombre,

            u.correo

        FROM docentes d

        INNER JOIN usuarios u

            ON d.id_usuario = u.id_usuario

        WHERE d.id_docente=%s

    """, (id,))

    docente = cursor.fetchone()

    # =============================
    # DISPONIBILIDAD
    # =============================

    cursor.execute("""

        SELECT

            lunes,

            martes,

            miercoles,

            jueves,

            viernes

        FROM disponibilidad_docente

        WHERE id_docente=%s

    """, (id,))

    disponibilidad = cursor.fetchone()

    conexion.close()

    return render_template(

        "admin/editar_docente.html",

        docente=docente,

        disponibilidad=disponibilidad

    )

# =====================================
# GESTIÓN DE USUARIOS
# =====================================

@app.route("/usuarios", methods=["GET", "POST"])
def usuarios():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        nombre = request.form["nombre"]
        correo = request.form["correo"]
        password = request.form["password"]
        rol = request.form["rol"]
        activo = request.form["activo"]

        cursor.execute("""
            INSERT INTO usuarios
            (
                nombre,
                correo,
                password,
                id_rol,
                activo
            )
            VALUES
            (%s,%s,%s,%s,%s)
        """,
        (
            nombre,
            correo,
            password,
            rol,
            activo
        ))

        # Obtener el ID del usuario recién creado
        id_usuario = cursor.lastrowid

        # Verificar si el usuario es Docente
        cursor.execute("""
            SELECT nombre
            FROM roles
            WHERE id_rol=%s
        """,(rol,))

        nombre_rol = cursor.fetchone()

        if nombre_rol and nombre_rol[0] == "Docente":

            cursor.execute("""
                INSERT INTO docentes
                (
                    id_usuario,
                    numero_empleado,
                    turno,
                    activo
                )
                VALUES
                (%s,'','Ambos',1)
            """,(id_usuario,))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Usuarios",

            "Crear Usuario",

            f"Registró el usuario '{nombre}'."

        )

        conexion.commit()

    cursor.execute("""
        SELECT
            u.id_usuario,
            u.nombre,
            u.correo,
            r.nombre,
            u.activo
        FROM usuarios u
        INNER JOIN roles r
            ON u.id_rol = r.id_rol
        ORDER BY u.id_usuario DESC
    """)

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT
            id_rol,
            nombre
        FROM roles
    """)

    roles = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/usuarios.html",
        usuarios=usuarios,
        roles=roles
    )


# =====================================
# EDITAR USUARIO
# =====================================

@app.route("/editar_usuario/<int:id>", methods=["GET","POST"])
def editar_usuario(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        nombre = request.form["nombre"]
        correo = request.form["correo"]
        activo = request.form["activo"]

        cursor.execute("""
            UPDATE usuarios
            SET
                nombre=%s,
                correo=%s,
                activo=%s
            WHERE id_usuario=%s
        """,
        (
            nombre,
            correo,
            activo,
            id
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Usuarios",

            "Editar Usuario",

            f"Actualizó el usuario '{nombre}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/usuarios")

    cursor.execute("""
        SELECT
            id_usuario,
            nombre,
            correo,
            id_rol,
            activo
        FROM usuarios
        WHERE id_usuario=%s
    """,(id,))

    usuario = cursor.fetchone()

    conexion.close()

    return render_template(
        "admin/editar_usuario.html",
        usuario=usuario
    )


@app.route("/logout")
def logout():

    session.clear()

    return redirect("/")

# =====================================
# GESTIÓN DE MATERIAS
# =====================================

@app.route("/materias", methods=["GET", "POST"])
def materias():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        clave = request.form["clave"]
        nombre = request.form["nombre"]
        horas_teoricas = request.form["horas_teoricas"]
        horas_practicas = request.form["horas_practicas"]
        semestre = request.form["semestre"]

        cursor.execute("""
            INSERT INTO materias
            (
                clave,
                nombre,
                horas_teoricas,
                horas_practicas,
                semestre
            )
            VALUES
            (%s,%s,%s,%s,%s)
        """,
        (
            clave,
            nombre,
            horas_teoricas,
            horas_practicas,
            semestre
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Materias",

            "Registrar Materia",

            f"Registró la materia '{nombre}'."

        )

        conexion.commit()

    cursor.execute("""
        SELECT
            id_materia,
            clave,
            nombre,
            horas_teoricas,
            horas_practicas,
            semestre
        FROM materias
        ORDER BY id_materia DESC
    """)

    materias = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/materias.html",
        materias=materias
    )


# =====================================
# EDITAR MATERIA
# =====================================

@app.route("/editar_materia/<int:id>", methods=["GET", "POST"])
def editar_materia(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        clave = request.form["clave"]
        nombre = request.form["nombre"]
        horas_teoricas = request.form["horas_teoricas"]
        horas_practicas = request.form["horas_practicas"]
        semestre = request.form["semestre"]

        cursor.execute("""
            UPDATE materias
            SET
                clave=%s,
                nombre=%s,
                horas_teoricas=%s,
                horas_practicas=%s,
                semestre=%s
            WHERE id_materia=%s
        """,
        (
            clave,
            nombre,
            horas_teoricas,
            horas_practicas,
            semestre,
            id
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Materias",

            "Editar Materia",

            f"Actualizó la materia '{nombre}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/materias")

    cursor.execute("""
        SELECT
            id_materia,
            clave,
            nombre,
            horas_teoricas,
            horas_practicas,
            semestre
        FROM materias
        WHERE id_materia=%s
    """, (id,))

    materia = cursor.fetchone()

    conexion.close()

    return render_template(
        "admin/editar_materia.html",
        materia=materia
    )


# =====================================
# ELIMINAR MATERIA
# =====================================

@app.route("/eliminar_materia/<int:id>")
def eliminar_materia(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    # Obtener el nombre antes de eliminar
    cursor.execute("""
        SELECT nombre
        FROM materias
        WHERE id_materia=%s
    """, (id,))

    materia = cursor.fetchone()

    cursor.execute("""
        DELETE FROM materias
        WHERE id_materia=%s
    """, (id,))

    # =============================
    # REGISTRAR HISTORIAL
    # =============================

    registrar_historial(

        conexion,

        session["usuario_id"],

        "Materias",

        "Eliminar Materia",

        f"Eliminó la materia '{materia[0]}'."

    )

    conexion.commit()

    conexion.close()

    return redirect("/materias")

# =====================================
# GESTIÓN DE GRUPOS
# =====================================

@app.route("/grupos", methods=["GET", "POST"])
def grupos():

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    if request.method == "POST":

        clave_grupo = request.form["clave_grupo"]
        semestre = request.form["semestre"]
        turno = request.form["turno"]
        cupo_maximo = request.form["cupo_maximo"]
        carrera = request.form["carrera"]

        cursor.execute("""

            INSERT INTO grupos
            (

                clave_grupo,
                semestre,
                turno,
                cupo_maximo,
                carrera,
                activo

            )

            VALUES
            (%s,%s,%s,%s,%s,%s)

        """,
        (

            clave_grupo,
            semestre,
            turno,
            cupo_maximo,
            carrera,
            1

        ))

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Grupos",

            "Registrar Grupo",

            f"Registró el grupo '{clave_grupo}'."

        )

        conexion.commit()

    cursor.execute("""

        SELECT

            id_grupo,
            clave_grupo,
            semestre,
            turno,
            cupo_maximo,
            carrera,
            activo

        FROM grupos

        ORDER BY semestre,
                clave_grupo

    """)

    grupos = cursor.fetchall()

    conexion.close()

    return render_template(

        "admin/grupos.html",

        grupos=grupos

    )

# =====================================
# EDITAR GRUPO
# =====================================

@app.route("/editar_grupo/<int:id>", methods=["GET", "POST"])
def editar_grupo(id):

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    if request.method == "POST":

        clave_grupo = request.form["clave_grupo"]
        semestre = request.form["semestre"]
        turno = request.form["turno"]
        cupo_maximo = request.form["cupo_maximo"]
        carrera = request.form["carrera"]

        cursor.execute("""

            UPDATE grupos

            SET

                clave_grupo=%s,
                semestre=%s,
                turno=%s,
                cupo_maximo=%s,
                carrera=%s

            WHERE id_grupo=%s

        """,
        (

            clave_grupo,
            semestre,
            turno,
            cupo_maximo,
            carrera,
            id

        ))

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Grupos",

            "Editar Grupo",

            f"Actualizó la información del grupo '{clave_grupo}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/grupos")

    cursor.execute("""

        SELECT *

        FROM grupos

        WHERE id_grupo=%s

    """, (id,))

    grupo = cursor.fetchone()

    conexion.close()

    return render_template(

        "admin/editar_grupo.html",

        grupo=grupo

    )

# =====================================
# ACTIVAR / DESACTIVAR GRUPO
# =====================================

@app.route("/cambiar_estado_grupo/<int:id>")
def cambiar_estado_grupo(id):

    conexion = pymysql.connect(

        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor

    )

    cursor = conexion.cursor()

    # Obtener datos actuales

    cursor.execute("""

        SELECT

            clave_grupo,
            activo

        FROM grupos

        WHERE id_grupo=%s

    """, (id,))

    grupo = cursor.fetchone()

    if grupo["activo"] == 1:

        nuevo_estado = 0
        accion = "Desactivar Grupo"
        descripcion = f"Desactivó el grupo '{grupo['clave_grupo']}'."

    else:

        nuevo_estado = 1
        accion = "Activar Grupo"
        descripcion = f"Activó el grupo '{grupo['clave_grupo']}'."

    cursor.execute("""

        UPDATE grupos

        SET activo=%s

        WHERE id_grupo=%s

    """,
    (
        nuevo_estado,
        id
    ))

    registrar_historial(

        conexion,

        session["usuario_id"],

        "Grupos",

        accion,

        descripcion

    )

    conexion.commit()

    conexion.close()

    return redirect("/grupos")

# =====================================
# GESTIÓN DE AULAS
# =====================================

@app.route("/aulas", methods=["GET", "POST"])
def aulas():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        edificio = request.form["edificio"]
        aula = request.form["aula"]
        capacidad = request.form["capacidad"]
        tipo = request.form["tipo"]

        cursor.execute("""
            INSERT INTO aulas
            (
                edificio,
                aula,
                capacidad,
                tipo
            )
            VALUES
            (%s,%s,%s,%s)
        """,
        (
            edificio,
            aula,
            capacidad,
            tipo
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Aulas",

            "Registrar Aula",

            f"Registró el aula '{edificio} - {aula}'."

        )

        conexion.commit()

    cursor.execute("""
        SELECT
            id_aula,
            edificio,
            aula,
            capacidad,
            tipo
        FROM aulas
        ORDER BY id_aula DESC
    """)

    aulas = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/aulas.html",
        aulas=aulas
    )


# =====================================
# EDITAR AULA
# =====================================

@app.route("/editar_aula/<int:id>", methods=["GET","POST"])
def editar_aula(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        edificio = request.form["edificio"]
        aula = request.form["aula"]
        capacidad = request.form["capacidad"]
        tipo = request.form["tipo"]

        cursor.execute("""
            UPDATE aulas
            SET
                edificio=%s,
                aula=%s,
                capacidad=%s,
                tipo=%s
            WHERE id_aula=%s
        """,
        (
            edificio,
            aula,
            capacidad,
            tipo,
            id
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Aulas",

            "Editar Aula",

            f"Actualizó el aula '{edificio} - {aula}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/aulas")

    cursor.execute("""
        SELECT
            id_aula,
            edificio,
            aula,
            capacidad,
            tipo
        FROM aulas
        WHERE id_aula=%s
    """,(id,))

    aula = cursor.fetchone()

    conexion.close()

    return render_template(
        "admin/editar_aula.html",
        aula=aula
    )


# =====================================
# ELIMINAR AULA
# =====================================

@app.route("/eliminar_aula/<int:id>")
def eliminar_aula(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    # Obtener la información antes de eliminar
    cursor.execute("""
        SELECT
            edificio,
            aula
        FROM aulas
        WHERE id_aula=%s
    """,(id,))

    datos = cursor.fetchone()

    cursor.execute("""
        DELETE FROM aulas
        WHERE id_aula=%s
    """,(id,))

    # =============================
    # REGISTRAR HISTORIAL
    # =============================

    registrar_historial(

        conexion,

        session["usuario_id"],

        "Aulas",

        "Eliminar Aula",

        f"Eliminó el aula '{datos[0]} - {datos[1]}'."

    )

    conexion.commit()

    conexion.close()

    return redirect("/aulas")

# =====================================
# CONFIGURACION ACADEMICA
# =====================================

@app.route("/configuracion_academica", methods=["GET", "POST"])
def configuracion_academica():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        clave = request.form["clave"]
        nombre = request.form["nombre"]
        fecha_inicio = request.form["fecha_inicio"]
        fecha_fin = request.form["fecha_fin"]

        cursor.execute("""
            INSERT INTO periodos_academicos
            (
                clave,
                nombre,
                fecha_inicio,
                fecha_fin
            )
            VALUES
            (%s,%s,%s,%s)
        """,
        (
            clave,
            nombre,
            fecha_inicio,
            fecha_fin
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Configuración Académica",

            "Crear Período",

            f"Registró el período académico '{nombre}'."

        )

        conexion.commit()

    cursor.execute("""
        SELECT
            id_periodo,
            clave,
            nombre,
            fecha_inicio,
            fecha_fin,
            estado
        FROM periodos_academicos
        ORDER BY id_periodo DESC
    """)

    periodos = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/configuracion_academica.html",
        periodos=periodos
    )


# =====================================
# EDITAR PERIODO
# =====================================

@app.route("/editar_periodo/<int:id>", methods=["GET","POST"])
def editar_configuracion(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    if request.method == "POST":

        clave = request.form["clave"]
        nombre = request.form["nombre"]
        fecha_inicio = request.form["fecha_inicio"]
        fecha_fin = request.form["fecha_fin"]

        cursor.execute("""
            UPDATE periodos_academicos
            SET
                clave=%s,
                nombre=%s,
                fecha_inicio=%s,
                fecha_fin=%s
            WHERE id_periodo=%s
        """,
        (
            clave,
            nombre,
            fecha_inicio,
            fecha_fin,
            id
        ))

        # =============================
        # REGISTRAR HISTORIAL
        # =============================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Configuración Académica",

            "Editar Período",

            f"Actualizó el período académico '{nombre}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/configuracion_academica")

    cursor.execute("""
        SELECT
            id_periodo,
            clave,
            nombre,
            fecha_inicio,
            fecha_fin,
            estado
        FROM periodos_academicos
        WHERE id_periodo=%s
    """,(id,))

    periodo = cursor.fetchone()

    conexion.close()

    return render_template(
        "admin/editar_configuracion.html",
        periodo=periodo
    )


# =====================================
# ACTIVAR PERIODO
# =====================================

@app.route("/activar_periodo/<int:id>")
def activar_periodo(id):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor()

    cursor.execute("""
        UPDATE periodos_academicos
        SET estado='Inactivo'
    """)

    cursor.execute("""
        UPDATE periodos_academicos
        SET estado='Activo'
        WHERE id_periodo=%s
    """,(id,))

    # Obtener nombre del periodo
    cursor.execute("""
        SELECT nombre
        FROM periodos_academicos
        WHERE id_periodo=%s
    """,(id,))

    periodo = cursor.fetchone()

    # =============================
    # REGISTRAR HISTORIAL
    # =============================

    registrar_historial(

        conexion,

        session["usuario_id"],

        "Configuración Académica",

        "Activar Período",

        f"Activó el período académico '{periodo[0]}'."

    )

    conexion.commit()

    conexion.close()

    return redirect("/configuracion_academica")

# =====================================
# CONFIGURAR CALENDARIO ACADÉMICO
# =====================================

@app.route("/configurar_calendario/<int:id_periodo>", methods=["GET", "POST"])
def configurar_calendario(id_periodo):

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

    cursor = conexion.cursor(pymysql.cursors.DictCursor)

    if request.method == "POST":

        hora_inicio_matutino = request.form["hora_inicio_matutino"]
        hora_fin_matutino = request.form["hora_fin_matutino"]

        hora_inicio_vespertino = request.form["hora_inicio_vespertino"]
        hora_fin_vespertino = request.form["hora_fin_vespertino"]

        duracion_bloque = request.form["duracion_bloque"]

        lunes = 1 if "lunes" in request.form else 0
        martes = 1 if "martes" in request.form else 0
        miercoles = 1 if "miercoles" in request.form else 0
        jueves = 1 if "jueves" in request.form else 0
        viernes = 1 if "viernes" in request.form else 0

        permitir = 1 if "permitir" in request.form else 0

        # ==========================================
        # CALENDARIO ACADÉMICO
        # ==========================================

        cursor.execute("""

            SELECT id_calendario

            FROM calendario_academico

            WHERE id_periodo=%s

        """,(id_periodo,))

        calendario = cursor.fetchone()

        if calendario:

            cursor.execute("""

                UPDATE calendario_academico

                SET

                    hora_inicio_matutino=%s,

                    hora_fin_matutino=%s,

                    hora_inicio_vespertino=%s,

                    hora_fin_vespertino=%s,

                    duracion_bloque=%s,

                    lunes=%s,

                    martes=%s,

                    miercoles=%s,

                    jueves=%s,

                    viernes=%s,

                    permitir_modificaciones_extra=%s

                WHERE id_periodo=%s

            """,
            (

                hora_inicio_matutino,

                hora_fin_matutino,

                hora_inicio_vespertino,

                hora_fin_vespertino,

                duracion_bloque,

                lunes,

                martes,

                miercoles,

                jueves,

                viernes,

                permitir,

                id_periodo

            ))

        else:

            cursor.execute("""

                INSERT INTO calendario_academico

                (

                    id_periodo,

                    hora_inicio_matutino,

                    hora_fin_matutino,

                    hora_inicio_vespertino,

                    hora_fin_vespertino,

                    duracion_bloque,

                    lunes,

                    martes,

                    miercoles,

                    jueves,

                    viernes,

                    permitir_modificaciones_extra

                )

                VALUES

                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)

            """,
            (

                id_periodo,

                hora_inicio_matutino,

                hora_fin_matutino,

                hora_inicio_vespertino,

                hora_fin_vespertino,

                duracion_bloque,

                lunes,

                martes,

                miercoles,

                jueves,

                viernes,

                permitir

            ))

        # ==========================================
        # PERIODO DE AJUSTES
        # ==========================================

        fecha_inicio = request.form["fecha_inicio_ajustes"]
        fecha_fin = request.form["fecha_fin_ajustes"]

        cursor.execute("""

            SELECT id_ajuste

            FROM periodos_ajustes

            WHERE id_periodo=%s

        """,(id_periodo,))

        ajuste = cursor.fetchone()

        if ajuste:

            cursor.execute("""

                UPDATE periodos_ajustes

                SET

                    fecha_inicio=%s,

                    fecha_fin=%s,

                    estado='Abierto'

                WHERE id_periodo=%s

            """,
            (

                fecha_inicio,

                fecha_fin,

                id_periodo

            ))

        else:

            cursor.execute("""

                INSERT INTO periodos_ajustes

                (

                    id_periodo,

                    fecha_inicio,

                    fecha_fin,

                    estado,

                    creado_por

                )

                VALUES

                (%s,%s,%s,%s,%s)

            """,
            (

                id_periodo,

                fecha_inicio,

                fecha_fin,

                "Abierto",

                session["usuario_id"]

            ))

        # ==========================================
        # OBTENER NOMBRE DEL PERIODO
        # ==========================================

        cursor.execute("""

            SELECT nombre

            FROM periodos_academicos

            WHERE id_periodo=%s

        """,(id_periodo,))

        periodo = cursor.fetchone()

        # ==========================================
        # REGISTRAR HISTORIAL
        # ==========================================

        registrar_historial(

            conexion,

            session["usuario_id"],

            "Configuración Académica",

            "Configurar Calendario",

            f"Actualizó la configuración del calendario del período '{periodo['nombre']}'."

        )

        conexion.commit()

        conexion.close()

        return redirect("/configuracion_academica")

    # ==========================================
    # CONSULTAS
    # ==========================================

    cursor.execute("""

        SELECT

            id_periodo,

            clave,

            nombre,

            fecha_inicio,

            fecha_fin,

            estado

        FROM periodos_academicos

        WHERE id_periodo=%s

    """,(id_periodo,))

    periodo = cursor.fetchone()

    cursor.execute("""

        SELECT

            hora_inicio_matutino,

            hora_fin_matutino,

            hora_inicio_vespertino,

            hora_fin_vespertino,

            duracion_bloque,

            lunes,

            martes,

            miercoles,

            jueves,

            viernes,

            permitir_modificaciones_extra

        FROM calendario_academico

        WHERE id_periodo=%s

    """,(id_periodo,))

    calendario = cursor.fetchone()

    if calendario:

        def formatear_hora(hora):

            if hora is None:
                return ""

            total_segundos = int(hora.total_seconds())

            horas = total_segundos // 3600
            minutos = (total_segundos % 3600) // 60

            return f"{horas:02d}:{minutos:02d}"

        calendario["hora_inicio_matutino"] = formatear_hora(calendario["hora_inicio_matutino"])
        calendario["hora_fin_matutino"] = formatear_hora(calendario["hora_fin_matutino"])
        calendario["hora_inicio_vespertino"] = formatear_hora(calendario["hora_inicio_vespertino"])
        calendario["hora_fin_vespertino"] = formatear_hora(calendario["hora_fin_vespertino"])

    cursor.execute("""

        SELECT

            fecha_inicio,

            fecha_fin,

            estado

        FROM periodos_ajustes

        WHERE id_periodo=%s

    """,(id_periodo,))

    ajustes = cursor.fetchone()

    if ajustes:

        if ajustes["fecha_inicio"] and not isinstance(ajustes["fecha_inicio"], str):
            ajustes["fecha_inicio"] = ajustes["fecha_inicio"].strftime("%Y-%m-%d")

        if ajustes["fecha_fin"] and not isinstance(ajustes["fecha_fin"], str):
            ajustes["fecha_fin"] = ajustes["fecha_fin"].strftime("%Y-%m-%d")

    conexion.close()

    return render_template(

        "admin/configurar_calendario.html",

        periodo=periodo,

        calendario=calendario,

        ajustes=ajustes

    )

# =====================================
# REPORTES Y AUDITORÍA
# =====================================

@app.route("/reportes")
def reportes():

    return render_template(
        "admin/reportes_inicio.html"
    )

# =====================================
# REPORTES DEL SISTEMA
# =====================================

@app.route("/reportes_sistema", methods=["GET"])
def reportes_sistema():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

    cursor = conexion.cursor()

    tipo = request.args.get("tipo", "")

    datos = []

    if tipo == "usuarios":

        cursor.execute("""

            SELECT

                u.nombre,
                u.correo,
                r.nombre AS rol,
                u.activo

            FROM usuarios u

            INNER JOIN roles r

                ON u.id_rol = r.id_rol

            ORDER BY u.nombre

        """)

        datos = cursor.fetchall()

    elif tipo == "docentes":

        cursor.execute("""

            SELECT

                d.numero_empleado,
                u.nombre,
                d.turno,
                d.activo

            FROM docentes d

            INNER JOIN usuarios u

                ON d.id_usuario = u.id_usuario

            ORDER BY u.nombre

        """)

        datos = cursor.fetchall()

    elif tipo == "materias":

        cursor.execute("""

            SELECT

                clave,
                nombre,
                horas_teoricas,
                horas_practicas,
                semestre

            FROM materias

            ORDER BY semestre, nombre

        """)

        datos = cursor.fetchall()

    elif tipo == "aulas":

        cursor.execute("""

            SELECT

                edificio,
                aula,
                capacidad,
                tipo

            FROM aulas

            ORDER BY edificio, aula

        """)

        datos = cursor.fetchall()

    conexion.close()

    return render_template(

        "admin/reportes.html",

        tipo=tipo,

        datos=datos

    )

# =====================================
# AUDITORÍA
# =====================================

@app.route("/auditoria", methods=["GET"])
def auditoria():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

    cursor = conexion.cursor()

    # =============================
    # CATÁLOGOS PARA FILTROS
    # =============================

    # =============================
    # USUARIOS CON ACTIVIDAD
    # =============================

    cursor.execute("""

        SELECT DISTINCT

            u.id_usuario,

            u.nombre

        FROM historial_cambios h

        INNER JOIN usuarios u

            ON h.id_usuario = u.id_usuario

        ORDER BY u.nombre

    """)

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT DISTINCT modulo
        FROM historial_cambios
        ORDER BY modulo
    """)

    modulos = cursor.fetchall()

    cursor.execute("""
        SELECT DISTINCT accion
        FROM historial_cambios
        ORDER BY accion
    """)

    acciones = cursor.fetchall()

    # =============================
    # RECIBIR FILTROS
    # =============================

    usuario = request.args.get("usuario", "")
    modulo = request.args.get("modulo", "")
    accion = request.args.get("accion", "")
    fecha_inicio = request.args.get("fecha_inicio", "")
    fecha_fin = request.args.get("fecha_fin", "")

    consulta = """

        SELECT

            h.id_historial,

            h.fecha,

            u.nombre,

            h.modulo,

            h.accion,

            h.descripcion

        FROM historial_cambios h

        INNER JOIN usuarios u

            ON h.id_usuario=u.id_usuario

        WHERE 1=1

    """

    parametros = []

    if usuario:

        consulta += " AND h.id_usuario=%s"

        parametros.append(usuario)

    if modulo:

        consulta += " AND h.modulo=%s"

        parametros.append(modulo)

    if accion:

        consulta += " AND h.accion=%s"

        parametros.append(accion)

    if fecha_inicio:

        consulta += " AND DATE(h.fecha)>= %s"

        parametros.append(fecha_inicio)

    if fecha_fin:

        consulta += " AND DATE(h.fecha)<= %s"

        parametros.append(fecha_fin)

    consulta += " ORDER BY h.fecha DESC"

    cursor.execute(consulta, parametros)

    historial = cursor.fetchall()

    conexion.close()

    return render_template(

        "admin/auditoria.html",

        historial=historial,

        usuarios=usuarios,

        modulos=modulos,

        acciones=acciones

    )

# =====================================
# EXPORTAR AUDITORÍA A EXCEL
# =====================================

@app.route("/exportar_excel")
def exportar_excel():

    conexion = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

    cursor = conexion.cursor()

    # =============================
    # FILTROS
    # =============================

    usuario = request.args.get("usuario", "")
    modulo = request.args.get("modulo", "")
    accion = request.args.get("accion", "")
    fecha_inicio = request.args.get("fecha_inicio", "")
    fecha_fin = request.args.get("fecha_fin", "")

    consulta = """

        SELECT

            h.fecha,

            u.nombre,

            h.modulo,

            h.accion,

            h.descripcion

        FROM historial_cambios h

        INNER JOIN usuarios u

            ON h.id_usuario=u.id_usuario

        WHERE 1=1

    """

    parametros = []

    if usuario:

        consulta += " AND h.id_usuario=%s"
        parametros.append(usuario)

    if modulo:

        consulta += " AND h.modulo=%s"
        parametros.append(modulo)

    if accion:

        consulta += " AND h.accion=%s"
        parametros.append(accion)

    if fecha_inicio:

        consulta += " AND DATE(h.fecha)>= %s"
        parametros.append(fecha_inicio)

    if fecha_fin:

        consulta += " AND DATE(h.fecha)<= %s"
        parametros.append(fecha_fin)

    consulta += " ORDER BY h.fecha DESC"

    cursor.execute(consulta, parametros)

    historial = cursor.fetchall()

    conexion.close()

    # =============================
    # CREAR EXCEL
    # =============================

    wb = Workbook()

    ws = wb.active

    ws.title = "Auditoría"

    encabezados = [

        "Fecha",
        "Usuario",
        "Módulo",
        "Acción",
        "Descripción"

    ]

    for columna, texto in enumerate(encabezados, start=1):

        celda = ws.cell(row=1, column=columna)

        celda.value = texto

        celda.font = Font(bold=True)

    fila = 2

    for h in historial:

        ws.cell(fila, 1).value = h["fecha"].strftime("%d/%m/%Y %H:%M")
        ws.cell(fila, 2).value = h["nombre"]
        ws.cell(fila, 3).value = h["modulo"]
        ws.cell(fila, 4).value = h["accion"]
        ws.cell(fila, 5).value = h["descripcion"]

        fila += 1

    archivo = io.BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return send_file(

        archivo,

        download_name="Historial_Auditoria.xlsx",

        as_attachment=True,

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

# =====================================
# IMPORTAR INFORMACIÓN
# =====================================

@app.route("/importar", methods=["GET", "POST"])
def importar():

    mensaje = ""
    error = False

    encontrados = 0
    validos = 0
    errores = 0
    importados = 0

    lista_errores = []
    materias_validas = []

    if request.method == "POST":

        tipo = request.form["tipo"]

        archivo = request.files["archivo"]

        # =====================================
        # CONEXIÓN MYSQL
        # =====================================

        conexion = pymysql.connect(

            host=DB_HOST,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME,
            cursorclass=pymysql.cursors.DictCursor

        )

        cursor = conexion.cursor()

        # =====================================
        # VALIDAR ARCHIVO
        # =====================================

        if archivo.filename == "":

            mensaje = "Debe seleccionar un archivo."

            error = True

        elif not archivo.filename.endswith(".xlsx"):

            mensaje = "El archivo debe tener formato .xlsx"

            error = True

        else:

            try:

                libro = load_workbook(archivo)

                hoja = libro.active

            except Exception:

                mensaje = "No fue posible leer el archivo Excel."

                error = True

                conexion.close()

                return render_template(

                    "admin/importar.html",

                    mensaje=mensaje,
                    error=error,
                    encontrados=encontrados,
                    validos=validos,
                    errores=errores,
                    importados=importados,
                    lista_errores=lista_errores

                )

            # =====================================
            # LEER ENCABEZADOS
            # =====================================

            encabezados_excel = [

                str(celda.value).strip()

                for celda in hoja[1]

                if celda.value is not None

            ]

            # =====================================
            # PLANTILLAS
            # =====================================

            plantillas = {

                "materias":[

                    "Clave",
                    "Nombre",
                    "Horas Teóricas",
                    "Horas Prácticas",
                    "Semestre"

                ],

                "docentes":[

                    "Número Empleado",
                    "Nombre",
                    "Correo",
                    "Turno"

                ],

                "aulas":[

                    "Edificio",
                    "Aula",
                    "Capacidad",
                    "Tipo"

                ]

            }

            encabezados_esperados = plantillas.get(tipo, [])

            # =====================================
            # VALIDAR ENCABEZADOS
            # =====================================

            if encabezados_excel != encabezados_esperados:

                mensaje = "El formato del archivo no coincide con la plantilla."

                error = True

                for i in range(max(len(encabezados_excel),len(encabezados_esperados))):

                    esperado = encabezados_esperados[i] if i < len(encabezados_esperados) else "(Sin columna)"

                    recibido = encabezados_excel[i] if i < len(encabezados_excel) else "(Sin columna)"

                    if esperado != recibido:

                        lista_errores.append({

                            "fila":"Encabezado",

                            "error":f"Se esperaba '{esperado}' y se encontró '{recibido}'."

                        })

            else:

                mensaje = "Formato del archivo correcto."

                error = False

                # =====================================
                # VALIDAR REGISTROS
                # =====================================

                encontrados = 0
                validos = 0
                errores = 0
                importados = 0

                fila_excel = 2

                claves_archivo = []

                for fila in hoja.iter_rows(min_row=2, values_only=True):

                    # ===============================
                    # IGNORAR FILAS VACÍAS
                    # ===============================

                    if all(celda is None for celda in fila):
                        continue

                    encontrados += 1

                    # =====================================
                    # MATERIAS
                    # =====================================

                    if tipo == "materias":

                        clave = str(fila[0]).strip() if fila[0] else ""
                        nombre = str(fila[1]).strip() if fila[1] else ""

                        horas_teoricas = fila[2]
                        horas_practicas = fila[3]
                        semestre = fila[4]

                        registro_valido = True
                        errores_fila = []

                        # ===============================
                        # CLAVE
                        # ===============================

                        if clave == "":

                            errores_fila.append("La clave está vacía.")
                            registro_valido = False

                        elif clave in claves_archivo:

                            errores_fila.append("La clave está duplicada dentro del archivo.")
                            registro_valido = False

                        else:

                            claves_archivo.append(clave)

                            # ===============================
                            # VALIDAR EN MYSQL
                            # ===============================

                            cursor.execute("""

                                SELECT id_materia

                                FROM materias

                                WHERE clave=%s

                            """,(clave,))

                            if cursor.fetchone():

                                errores_fila.append(
                                    "La clave ya existe en el sistema."
                                )

                                registro_valido = False

                        # ===============================
                        # NOMBRE
                        # ===============================

                        if nombre == "":

                            errores_fila.append(
                                "El nombre está vacío."
                            )

                            registro_valido = False

                        # ===============================
                        # HORAS TEÓRICAS
                        # ===============================

                        try:

                            horas_teoricas = int(horas_teoricas)

                            if horas_teoricas < 0:

                                raise ValueError

                        except:

                            errores_fila.append(
                                "Horas teóricas inválidas."
                            )

                            registro_valido = False

                        # ===============================
                        # HORAS PRÁCTICAS
                        # ===============================

                        try:

                            horas_practicas = int(horas_practicas)

                            if horas_practicas < 0:

                                raise ValueError

                        except:

                            errores_fila.append(
                                "Horas prácticas inválidas."
                            )

                            registro_valido = False

                        # ===============================
                        # SEMESTRE
                        # ===============================

                        try:

                            semestre = int(semestre)

                            if semestre < 1 or semestre > 9:

                                raise ValueError

                        except:

                            errores_fila.append(
                                "Semestre inválido."
                            )

                            registro_valido = False

                        # ===============================
                        # CONTADORES
                        # ===============================

                        if registro_valido:

                            validos += 1

                            materias_validas.append({

                                "clave": clave,

                                "nombre": nombre,

                                "horas_teoricas": horas_teoricas,

                                "horas_practicas": horas_practicas,

                                "semestre": semestre

                            })

                        else:

                            errores += 1

                            lista_errores.append({

                                "fila": fila_excel,

                                "error": " | ".join(errores_fila)

                            })

                    # =====================================
                    # DOCENTES
                    # (Lo desarrollaremos después)
                    # =====================================

                    elif tipo == "docentes":

                        validos += 1

                    # =====================================
                    # AULAS
                    # (Lo desarrollaremos después)
                    # =====================================

                    elif tipo == "aulas":

                        validos += 1

                    fila_excel += 1

                conexion.close()

                # =====================================
                # IMPORTAR MATERIAS
                # =====================================

                accion = request.form.get("accion")

                if accion == "importar" and tipo == "materias":

                    conexion = pymysql.connect(

                        host=DB_HOST,
                        user=DB_USER,
                        password=DB_PASSWORD,
                        database=DB_NAME,
                        cursorclass=pymysql.cursors.DictCursor

                    )

                    cursor = conexion.cursor()

                    importados = 0

                    for materia in materias_validas:

                        cursor.execute("""

                            INSERT INTO materias
                            (

                                clave,
                                nombre,
                                horas_teoricas,
                                horas_practicas,
                                semestre

                            )

                            VALUES
                            (%s,%s,%s,%s,%s)

                        """,
                        (

                            materia["clave"],
                            materia["nombre"],
                            materia["horas_teoricas"],
                            materia["horas_practicas"],
                            materia["semestre"]

                        ))

                        importados += 1

                    conexion.commit()

                    # =====================================
                    # HISTORIAL
                    # =====================================

                    registrar_historial(

                        conexion,

                        session["usuario_id"],

                        "Importación",

                        "Importar",

                        f"Importó {importados} materias desde archivo Excel."

                    )

                    conexion.close()

                    mensaje = f"Se importaron correctamente {importados} materias."

                    error = False

                

    return render_template(

    "admin/importar.html",

    mensaje=mensaje,

    error=error,

    encontrados=encontrados,

    validos=validos,

    errores=errores,

    importados=importados,

    lista_errores=lista_errores

)

# =====================================
# INICIO DE LA APP
# =====================================

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )
